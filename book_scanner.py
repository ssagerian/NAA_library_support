import openpyxl
from tkinter import filedialog, Tk

def scan_books(excel_path, scanned_books):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    all_books = set(str(sheet.cell(row=i, column=1).value) for i in range(2, sheet.max_row + 1))
    not_found_books = set(scanned_books) - all_books

    return list(not_found_books)

def main():
    root = Tk()
    root.withdraw()
    excel_path = filedialog.askopenfilename(title="Select Excel file")

    scanned_books = input("Scan books (comma-separated): ").split(',')

    not_found_books = scan_books(excel_path, scanned_books)

    if not_found_books:
        print("Books not found:")
        for book in not_found_books:
            print(book)
    else:
        print("All books found!")

if __name__ == "__main__":
    main()
