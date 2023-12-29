import openpyxl
import json

def save_to_file(book_data):
    with open("books_seen.json", "w") as file:
        json.dump(book_data, file, indent=2)
    print("Data saved to 'books_seen.json'.")

def read_excel_to_list(file_path, max_rows_per_sheet=None):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Initialize the dictionary to store data for each sheet
    workbook_data = {}

    # Iterate through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        # Select the current sheet
        sheet = workbook[sheet_name]

        # Get the header row
        headers = [cell.value for cell in sheet[1]]

        # Initialize the list to store dictionaries for the current sheet
        sheet_data = []

        # Iterate over rows starting from the second row (index 2)
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Create a dictionary using the headers and current row values
            data_dict = dict(zip(headers, row))
            sheet_data.append(data_dict)

            # Check if a maximum number of rows is specified and break if reached
            if max_rows_per_sheet is not None and len(sheet_data) >= max_rows_per_sheet:
                break

            # Check for an empty row (assuming that an empty first cell indicates an empty row)
            if row[0] is None:
                break

        # Store the data for the current sheet in the dictionary
        workbook_data[sheet_name] = sheet_data

    return workbook_data

if __name__ == "__main__":
    # Replace 'path/to/your/excel/file.xlsx' with your actual file path
    excel_file_path = 'compiled.xlsx'

    # Specify the maximum number of rows to process per sheet (set to None to process all rows)
    max_rows_per_sheet = 1000

    # Get data from all sheets in the Excel workbook
    all_sheets_data = read_excel_to_list(excel_file_path, max_rows_per_sheet)

    save_to_file(all_sheets_data)

    # Print the resulting dictionary containing data for all sheets
    for sheet_name, sheet_data in all_sheets_data.items():
        print(f"Sheet: {sheet_name}")
        print(sheet_data)
        print()
